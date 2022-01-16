#  -*-  coding:utf-8 -*-
from openpyxl import load_workbook, Workbook

file = "/Users/le/Documents/pyproject/lottery/utils/抽奖员工.xlsx"


class ExcelUtils:
    def __init__(self):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb['Sheet']

    def excel_read(self):
        lists = []
        print(self.ws.max_row)
        # 遍历excel中数据
        for i in range(1, self.ws.max_row + 1):
            name = self.ws.cell(i, 1).value
            lists.append(name)
        return lists

    # 通过值获取行数
    def get_cell_row(self, cell_value):
        lists = []
        for i in range(1, self.ws.max_row + 1):
            name = self.ws.cell(i, 1).value
            for value in cell_value:
                if name == value:
                    lists.append(i)
        return lists

    # 删除一行
    def delete_row(self, which_rows):
        for i in which_rows:
            self.ws.delete_rows(int(i))
            self.wb.save(self.file)


def excel_write(names):
    wb = Workbook()
    ws = wb['Sheet']
    for i in range(1, len(names) + 1):
        ws.cell(i, 1, names[i - 1])
    try:
        wb.save("抽奖员工.xlsx")
    except Exception as e:
        print("error :{}".format(e))


# 模拟员工数据
if __name__ == "__main__":
    from faker import Faker

    f = Faker(locale='zh_CN')
    name_list = []
    for i in range(100):
        name_list.append(f.name())
    for i in range(100):
        excel_write(name_list)

